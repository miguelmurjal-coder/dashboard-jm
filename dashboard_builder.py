from __future__ import annotations

import calendar
import csv
import io
import json
import re
import sys
import urllib.request
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


VERSION = "1.9.1"
VACATION_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQXk2yWBJ5SFJUPhJG7oBWhqs5tJylVsDWBl6GGndu2oWrwti6e6csHZpmxaJG9ywzStdmR0_4Q2URX/pub?gid=1810617663&single=true&output=csv"
DEFAULT_SOURCE = Path(r"C:\Users\pke\Desktop\JM_NZ_2023\INDICADORES OPERACIONAIS\OM2026\GESTÃO\INDICADOR MASTER 2026.xlsx")
HERE = Path(__file__).resolve().parent
OUTPUT = HERE / "index.html"
TEMPLATE = OUTPUT


def number(value):
    return float(value) if isinstance(value, (int, float)) else 0.0


def text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def date_time_parts(value):
    if isinstance(value, datetime):
        return value.date().isoformat(), value.strftime("%H:%M")
    s = text(value).replace("T", " ")
    for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            d = datetime.strptime(s, fmt)
            return d.date().isoformat(), d.strftime("%H:%M") if "%H" in fmt else ""
        except ValueError:
            pass
    return iso_date(value), ""


def extract_invoices(ws):
    # A grelha detalhada da FRC (M:AM) inclui FRC e CRDFC, assim como a
    # classificação manual PS/FO/ADB/FS. Agregamos por documento para que
    # documentos com várias linhas de artigos apareçam apenas uma vez.
    documents = {}
    for row in ws.iter_rows(min_row=7, min_col=2, max_col=39, values_only=True):
        series = text(row[20]).upper()
        number_value = row[21]
        inv_date = iso_date(row[11])
        if series not in {"FRC", "CRDFC"} or number_value is None or not inv_date:
            continue
        key = (series, text(number_value))
        if key not in documents:
            documents[key] = {
                "date": inv_date, "series": series, "number": number_value,
                "ref": row[12], "client": text(row[15]), "nif": text(row[16]),
                "net": 0.0, "vat": 0.0, "total": 0.0,
                "ps": 0.0, "fo": 0.0, "adb": 0.0, "fs": 0.0,
                "articles": [],
            }
        doc = documents[key]
        line_net, line_gross = number(row[32]), number(row[33])
        # No ficheiro, CRDFC já surge com sinal negativo na grelha FRC.
        doc["net"] += line_net
        doc["total"] += line_gross
        doc["ps"] += number(row[34])
        doc["fo"] += number(row[35])
        doc["adb"] += number(row[36])
        doc["fs"] += number(row[37])
        article_name = text(row[27])
        if article_name:
            doc["articles"].append({
                "article": article_name, "code": text(row[26]),
                "qty": number(row[28]), "net": number(row[30]),
                "gross": number(row[31]),
            })
    invoices = []
    for doc in documents.values():
        for field in ("net", "total", "ps", "fo", "adb", "fs"):
            doc[field] = round(doc[field], 2)
        doc["vat"] = round(doc["total"] - doc["net"], 2)
        invoices.append(doc)
    return sorted(invoices, key=lambda x: (x["date"], x["series"], text(x["number"])))


def extract_credits(ws):
    credits = []
    for row in ws.iter_rows(min_row=7, min_col=3, max_col=19, values_only=True):
        d = iso_date(row[1])
        if not d or not text(row[2]) or row[3] is None:
            continue
        order = row[11]
        order_text = text(order)
        credits.append({
            "monthtag": text(row[0]), "date": d, "series": text(row[2]),
            "number": row[3], "ref": row[4], "client": text(row[5]), "nif": text(row[6]),
            "net": number(row[7]), "vat": number(row[8]), "total": number(row[9]),
            "obs": text(row[10]), "order": order,
            "orderUrl": f"https://www.pke.pt/admin/orders/{order_text}/edit#tab-form-4" if order_text else "",
            "ps": number(row[13]), "fo": number(row[14]), "adb": number(row[15]), "fs": number(row[16]),
        })
    return credits


def extract_daily(ws):
    found = {}
    rows = list(ws.iter_rows(min_row=1, min_col=42, max_col=43, values_only=True))
    for i, row in enumerate(rows[:-2]):
        d = iso_date(row[0])
        if d and d.startswith("2026-"):
            net, gross = rows[i + 2]
            if isinstance(net, (int, float)) and isinstance(gross, (int, float)):
                found[d] = {"date": d, "net": round(float(net), 2), "gross": round(float(gross), 2)}
    if not found:
        raise RuntimeError("Não foram encontrados os resumos diários na folha FRC.")
    last = max(datetime.fromisoformat(x).date() for x in found)
    daily = {}
    monthly_net, monthly_gross = [], []
    for month in range(1, last.month + 1):
        items = []
        days = calendar.monthrange(2026, month)[1]
        for day in range(1, days + 1):
            key = date(2026, month, day).isoformat()
            items.append(found.get(key, {"date": key, "net": 0.0, "gross": 0.0}))
        daily[str(month)] = items
        monthly_net.append(round(sum(x["net"] for x in items), 2))
        monthly_gross.append(round(sum(x["gross"] for x in items), 2))
    return daily, monthly_net, monthly_gross


def extract_backoffice_daily(ws):
    rows = []
    for row in ws.iter_rows(min_row=48, min_col=5, max_col=20, values_only=True):
        d = iso_date(row[0])
        if not d or not d.startswith("2026-"):
            continue
        rows.append({
            "date": d,
            "psPaid": number(row[7]), "psUnpaid": number(row[8]),
            "foPaid": number(row[9]), "foUnpaid": number(row[10]),
            "adbPaid": number(row[11]), "adbUnpaid": number(row[12]),
            "fsPaid": number(row[13]), "fsUnpaid": number(row[14]),
            "total": number(row[15]),
        })
    if not rows:
        raise RuntimeError("Não foram encontrados os totais do Back Office em Overview2026!E48:T.")
    return rows


def extract_orders(workbook):
    sheet_months = {
        "BO | 2026": {1, 2}, "BO | 2026 (new)": {3, 4}, "BO |2026 MAI": {5},
        "BO |2026 JUN": {6}, "BO |2026 JUL": {7}, "BO |2026 AGT": {8},
        "BO |2026 SET": {9}, "BO |2026 OUT": {10}, "BO |2026 NOV e DEZ": {11, 12},
    }
    orders, seen = [], set()
    for sheet_name, allowed_months in sheet_months.items():
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        active = None
        for row in ws.iter_rows(min_row=4, min_col=1, max_col=51, values_only=True):
            oid = row[29]
            if isinstance(oid, (int, float)) and int(oid) > 1000:
                d, tm = date_time_parts(row[3])
                if not d or int(d[5:7]) not in allowed_months:
                    active = None
                    continue
                oid = int(oid)
                if oid in seen:
                    active = None
                    continue
                seen.add(oid)
                active = {
                    "id": oid, "date": d, "time": tm, "scheduledDate": iso_date(row[2]),
                    "type": text(row[4]), "client": text(row[5]), "email": text(row[6]),
                    "phone": text(row[7]), "paymentStatus": text(row[35]),
                    "gross": number(row[36]), "deposit": number(row[37]), "voucher": text(row[11]),
                    "discount": text(row[12]), "status": text(row[32] or row[13]),
                    "contactType": text(row[14]), "author": text(row[33] or row[15]),
                    "products": [], "net": number(row[38]),
                    "ps": number(row[39]), "psUnpaid": number(row[40]),
                    "fo": number(row[41]), "foUnpaid": number(row[42]),
                    "adb": number(row[43]), "adbUnpaid": number(row[44]),
                    "fs": number(row[45]), "fsUnpaid": number(row[46]),
                    "order": text(row[47] or oid), "easypay": text(row[48]), "fr": text(row[49]),
                    "payment": number(row[50]), "invoice": text(row[49]),
                    "url": f"https://www.pke.pt/admin/orders/{oid}/edit#tab-form-4",
                }
                product = text(row[34] or row[19])
                if product:
                    active["products"].append(product)
                orders.append(active)
            elif active:
                product = text(row[34] or row[19])
                if product and product not in active["products"]:
                    active["products"].append(product)
    orders.sort(key=lambda x: (x["date"], x["time"], x["id"]))
    return orders


def replace_json_assignment(html, variable, payload):
    marker = f"const {variable}="
    start = html.index(marker) + len(marker)
    decoder = json.JSONDecoder()
    _, length = decoder.raw_decode(html[start:])
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), allow_nan=False)
    return html[:start] + encoded + html[start + length:]


def extract_hr(ws):
    """Extrai apenas os campos operacionais necessários ao separador RH."""
    people = []
    for row in ws.iter_rows(min_row=21, max_col=32, values_only=True):
        active = text(row[2]).upper()
        name = text(row[6])
        if active not in {"SIM", "NÃO", "NAO"} or not name:
            continue
        people.append({
            "name": " ".join(part.capitalize() for part in name.split()),
            "status": "Ativo" if active == "SIM" else "Inativo",
            "department": text(row[3]),
            "role": text(row[7]),
            "schedule": "Tempo parcial" if text(row[5]).lower() == "x" else "Tempo completo",
            "contract": text(row[27]),
            "entryDate": iso_date(row[30]),
            "exitDate": iso_date(row[31]),
        })
    people.sort(key=lambda item: (item["department"], item["name"]))
    return {"sheet": "RH", "people": people}


def extract_vacations():
    request = urllib.request.Request(VACATION_URL, headers={"User-Agent": "PKE-Dashboard/1.0"})
    with urllib.request.urlopen(request, timeout=30) as response:
        rows = list(csv.reader(io.StringIO(response.read().decode("utf-8-sig"))))
    people, team_by_name, current_team = [], {}, ""
    for raw in rows[3:30]:
        row = raw + [""] * (20 - len(raw))
        name, team = row[3].strip(), row[2].strip()
        if team:
            current_team = team
        if not name or name.upper().startswith(("FERIADO", "PLANO")):
            continue
        try:
            allowance = float(row[5] or 0)
            monthly = [float(row[i] or 0) for i in range(6, 18)]
            used, remaining = float(row[18] or 0), float(row[19] or 0)
        except ValueError:
            continue
        person = {"code": row[0].strip(), "name": name, "team": team or current_team,
                  "allowance": allowance, "monthly": monthly, "used": used, "remaining": remaining,
                  "incompatible": [x.strip() for x in row[1].split(",") if x.strip()]}
        people.append(person)
        team_by_name[name] = person["team"]
    entries, date_pattern = [], re.compile(r"(\d{2}/\d{2}/\d{4})$")
    for index, row in enumerate(rows):
        if len(row) < 5 or not date_pattern.search(row[4] or ""):
            continue
        dates = []
        for col in range(4, len(row)):
            match = date_pattern.search(row[col] or "")
            if not match:
                break
            dates.append((col, datetime.strptime(match.group(1), "%d/%m/%Y").date().isoformat()))
        for raw_person in rows[index + 2:index + 22]:
            person_row = raw_person + [""] * max(0, max((c for c, _ in dates), default=3) + 1 - len(raw_person))
            name = person_row[3].strip() if len(person_row) > 3 else ""
            if not name or name.upper().startswith(("PLANO", "TOTAL")):
                continue
            for col, day in dates:
                value = person_row[col].strip() if col < len(person_row) else ""
                if not value:
                    continue
                try:
                    amount = float(value.replace(",", "."))
                except ValueError:
                    amount = 1.0
                if amount > 0:
                    entries.append({"date": day, "name": name, "code": person_row[0].strip(),
                                    "team": team_by_name.get(name, person_row[2].strip()),
                                    "amount": amount, "type": "Férias"})
    holidays, holiday_pattern = [], re.compile(r"^(\d{1,2}) de ([^\s]+).+: (.+)$")
    month_map = {"janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4, "maio": 5, "junho": 6,
                 "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12}
    for row in rows[:30]:
        for cell in row:
            match = holiday_pattern.match(cell.strip())
            if match and match.group(2).lower() in month_map:
                holidays.append({"date": date(2026, month_map[match.group(2).lower()], int(match.group(1))).isoformat(), "name": match.group(3).strip()})
    edited = ""
    if rows and "Última edição:" in ",".join(rows[0]):
        edited = ",".join(rows[0]).split("Última edição:", 1)[1].strip()
        edited = re.sub(r"\s+por\s+.*$", "", edited).strip(" ,")
    return {"source": VACATION_URL, "edited": edited, "people": people, "entries": entries, "holidays": holidays}


def build(source):
    if not source.exists():
        raise FileNotFoundError(f"Ficheiro fonte não encontrado: {source}")
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template não encontrado: {TEMPLATE}")
    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    try:
        invoices = extract_invoices(workbook["FRC"])
        credits = extract_credits(workbook["NC"])
        daily, monthly_net, monthly_gross = extract_daily(workbook["FRC"])
        backoffice_daily = extract_backoffice_daily(workbook["Overview2026"])
        hr = extract_hr(workbook["RH"])
    finally:
        workbook.close()
    if len(invoices) < 100 or len(backoffice_daily) < 300:
        raise RuntimeError("A validação detetou poucos registos; o dashboard anterior foi mantido.")
    html = TEMPLATE.read_text(encoding="utf-8")
    html = replace_json_assignment(html, "DB", {"invoices": invoices, "credits": credits})
    html = replace_json_assignment(html, "OVERVIEW_SOURCE", {
        "monthlyNet": monthly_net, "monthlyGross": monthly_gross, "daily": daily,
        "orders": [], "backOfficeDaily": backoffice_daily,
    })
    html = replace_json_assignment(html, "RH_SOURCE", hr)
    vacations = None
    try:
        vacations = extract_vacations()
        html = replace_json_assignment(html, "VACATION_SOURCE", vacations)
    except Exception as exc:
        print(f"Aviso: plano de férias não atualizado ({exc}); foram mantidos os últimos dados.")
    html = re.sub(r"<title>.*?</title>", f"<title>PKE Automotive Dashboard v{VERSION}</title>", html, count=1)
    stamp = datetime.now().astimezone().strftime("%d/%m/%Y às %H:%M")
    update_marker = (
        "document.getElementById('today').innerHTML="
        "'<span>Dados atualizados</span><strong>" + stamp + "</strong>';"
    )
    html, marker_count = re.subn(
        r"document\.getElementById\('today'\)[^\n]*",
        update_marker,
        html,
        count=1,
    )
    if marker_count != 1:
        raise RuntimeError("Não foi possível atualizar o marcador de data do dashboard.")
    temp = OUTPUT.with_suffix(".novo.html")
    temp.write_text(html, encoding="utf-8")
    temp.replace(OUTPUT)
    archive = HERE.parent / "PKE-Automotive-Dashboard.zip"
    archive_temp = archive.with_suffix(".novo.zip")
    with zipfile.ZipFile(archive_temp, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as package:
        for item in sorted(HERE.iterdir()):
            if item.is_file() and not item.name.endswith(".pyc"):
                package.write(item, f"{HERE.name}/{item.name}")
    archive_temp.replace(archive)
    print(f"Dashboard v{VERSION} atualizado com sucesso.")
    print(f"Faturas: {len(invoices)} | Notas de crédito: {len(credits)} | Dias Back Office: {len(backoffice_daily)}")
    print(f"Dados até: {max(x['date'] for x in invoices + credits)}")
    if vacations:
        print(f"Férias: {len(vacations['people'])} colaboradores | {len(vacations['entries'])} marcações")
    print(f"ZIP atualizado: {archive.name}")


if __name__ == "__main__":
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    build(source)
